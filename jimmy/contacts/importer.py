"""Import contacts from FIJI alumni Excel, LinkedIn CSV, macOS Contacts, or manual additions."""
import csv
import re
from pathlib import Path
from typing import Optional

from .db import add_contact, _conn, _ensure_affiliation


def import_fiji_excel(path: str) -> int:
    """Import FIJI alumni dataset from the multi-sheet Excel workbook.
    Reads 'Tier 1 Network' (enriched) and 'Master Table' (broad) sheets.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    count = 0
    seen = set()

    # Import Tier 1 Network first (best data)
    if "Tier 1 Network" in wb.sheetnames:
        count += _import_fiji_sheet(wb["Tier 1 Network"], seen, tier_default="1")

    # Import from Tier 1-2 Deep Dives (has tier column)
    if "Tier 1-2 Deep Dives" in wb.sheetnames:
        count += _import_fiji_deep_dives(wb["Tier 1-2 Deep Dives"], seen)

    # Import Master Table for broader coverage (skip deceased)
    if "Master Table" in wb.sheetnames:
        count += _import_fiji_master(wb["Master Table"], seen)

    wb.close()
    return count


def _import_fiji_sheet(ws, seen: set, tier_default: str = "") -> int:
    """Import from 'Tier 1 Network' sheet.
    Columns: Last Name, First Name, Class, Industry, Sub-Industry, Current Company, Role/Title, Location
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col = {h: i for i, h in enumerate(header)}

    def g(row, *keys):
        for k in keys:
            if k in col and col[k] < len(row) and row[col[k]]:
                return str(row[col[k]]).strip()
        return ""

    count = 0
    for row in rows[1:]:
        if not any(row):
            continue
        first = g(row, "first name", "first")
        last = g(row, "last name", "last")
        if not first and not last:
            continue

        key = f"{first.lower()}|{last.lower()}"
        if key in seen:
            continue
        seen.add(key)

        location = g(row, "location")
        city, region = _parse_location(location)

        role = g(row, "role/title", "role", "title")
        # Extract city from role if it has (NYC) etc
        if not location and role:
            m = re.search(r'\(([^)]+)\)', role)
            if m:
                city, region = _parse_location(m.group(1))

        data = {
            "first_name": first,
            "last_name": last,
            "company": g(row, "current company", "company"),
            "role": role,
            "industry": g(row, "industry"),
            "sub_industry": g(row, "sub-industry", "sub_industry"),
            "city": city,
            "region": region,
            "tier": tier_default,
            "source": "fiji_alumni",
            "affiliations": ["Columbia", "FIJI"],
        }

        class_year = g(row, "class", "class year")
        if class_year:
            data["notes"] = f"Class of {class_year}"

        add_contact(data)
        count += 1
    return count


def _import_fiji_deep_dives(ws, seen: set) -> int:
    """Import from 'Tier 1-2 Deep Dives' sheet.
    Has header on row 3: Last Name, First Name, Class Year, Status, Tier, Industry, Sub-Industry, Current Company, Role/Title, ...
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains 'Last Name')
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(str(c).strip().lower() == "last name" for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        return 0

    header = [str(c).strip().lower() if c else "" for c in rows[header_idx]]
    col = {h: i for i, h in enumerate(header)}

    def g(row, *keys):
        for k in keys:
            if k in col and col[k] < len(row) and row[col[k]]:
                return str(row[col[k]]).strip()
        return ""

    count = 0
    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        first = g(row, "first name", "first")
        last = g(row, "last name", "last")
        if not first and not last:
            continue

        status = g(row, "status")
        if "deceased" in status.lower():
            continue

        key = f"{first.lower()}|{last.lower()}"
        if key in seen:
            continue
        seen.add(key)

        tier_raw = g(row, "tier")
        tier = re.sub(r'[^0-9]', '', tier_raw)[:1] if tier_raw else ""

        location = g(row, "location")
        city, region = _parse_location(location)

        data = {
            "first_name": first,
            "last_name": last,
            "company": g(row, "current company", "company"),
            "role": g(row, "role/title", "role", "title"),
            "industry": g(row, "industry"),
            "sub_industry": g(row, "sub-industry"),
            "tier": tier,
            "city": city,
            "region": region,
            "source": "fiji_alumni",
            "affiliations": ["Columbia", "FIJI"],
        }

        class_year = g(row, "class year", "class")
        if class_year:
            data["notes"] = f"Class of {class_year}"

        add_contact(data)
        count += 1
    return count


def _import_fiji_master(ws, seen: set) -> int:
    """Import from 'Master Table' — broad list. Skip deceased and already-imported.
    Columns: Last Name, First Name, Middle Initial, Suffix, Class Year, Status, Birth Date, Age, ...
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col = {h: i for i, h in enumerate(header)}

    def g(row, *keys):
        for k in keys:
            if k in col and col[k] < len(row) and row[col[k]]:
                return str(row[col[k]]).strip()
        return ""

    count = 0
    for row in rows[1:]:
        if not any(row):
            continue
        first = g(row, "first name", "first")
        last = g(row, "last name", "last")
        if not first and not last:
            continue

        status = g(row, "status")
        if "deceased" in status.lower() or "unknown" in status.lower():
            continue

        key = f"{first.lower()}|{last.lower()}"
        if key in seen:
            continue
        seen.add(key)

        data = {
            "first_name": first,
            "last_name": last,
            "source": "fiji_alumni",
            "affiliations": ["Columbia", "FIJI"],
        }

        class_year = g(row, "class year", "class")
        if class_year:
            data["notes"] = f"Class of {class_year}"

        add_contact(data)
        count += 1
    return count


def import_linkedin_csv(path: str) -> int:
    """Import LinkedIn connections export CSV."""
    count = 0
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            first = (row.get("First Name") or "").strip()
            last = (row.get("Last Name") or "").strip()
            if not first and not last:
                continue
            connected = (row.get("Connected On") or "").strip()
            data = {
                "first_name": first,
                "last_name": last,
                "company": (row.get("Company") or "").strip(),
                "role": (row.get("Position") or "").strip(),
                "email": (row.get("Email Address") or "").strip(),
                "warmth_last_contact": _parse_date(connected),
                "source": "linkedin",
                "affiliations": [],
            }
            add_contact(data)
            count += 1
    return count


def import_macos_contacts() -> int:
    """Import contacts from macOS Contacts.app via AppleScript/contacts framework."""
    import subprocess
    import json

    script = '''
    tell application "Contacts"
        set output to "["
        set pList to every person
        repeat with p in pList
            set fn to first name of p as text
            set ln to last name of p as text
            if fn is missing value then set fn to ""
            if ln is missing value then set ln to ""
            if fn is "" and ln is "" then
            else
                set em to ""
                try
                    set em to value of first email of p as text
                end try
                set ph to ""
                try
                    set ph to value of first phone of p as text
                end try
                set co to ""
                try
                    set co to organization of p as text
                    if co is missing value then set co to ""
                end try
                set jt to ""
                try
                    set jt to job title of p as text
                    if jt is missing value then set jt to ""
                end try
                set nt to ""
                try
                    set nt to note of p as text
                    if nt is missing value then set nt to ""
                end try
                set output to output & "{\\"fn\\":\\"" & fn & "\\",\\"ln\\":\\"" & ln & "\\",\\"em\\":\\"" & em & "\\",\\"ph\\":\\"" & ph & "\\",\\"co\\":\\"" & co & "\\",\\"jt\\":\\"" & jt & "\\",\\"nt\\":\\"" & nt & "\\"},"
            end if
        end repeat
        if output ends with "," then set output to text 1 thru -2 of output
        set output to output & "]"
        return output
    end tell
    '''
    result = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        raise RuntimeError(f"AppleScript failed: {result.stderr}")

    raw = result.stdout.strip()
    # Clean up AppleScript JSON quirks
    raw = raw.replace('\r', ' ').replace('\n', ' ')
    try:
        people = json.loads(raw)
    except json.JSONDecodeError:
        # Fallback: try fixing common issues
        raw = re.sub(r'\\(?!["\\/bfnrtu])', r'\\\\', raw)
        people = json.loads(raw)

    count = 0
    for p in people:
        first = p.get("fn", "").strip()
        last = p.get("ln", "").strip()
        if not first and not last:
            continue
        data = {
            "first_name": first,
            "last_name": last,
            "email": p.get("em", "").strip(),
            "phone": p.get("ph", "").strip(),
            "company": p.get("co", "").strip(),
            "role": p.get("jt", "").strip(),
            "notes": p.get("nt", "").strip(),
            "source": "apple_contacts",
            "affiliations": [],
        }
        add_contact(data)
        count += 1
    return count


def import_gmail_contacts(max_messages: int = 500) -> int:
    """Extract unique contacts from Gmail sent/received messages across all Google accounts."""
    from ..config import GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET
    from ..ingestion.google_auth import get_all_credentials
    from googleapiclient.discovery import build
    import email.utils

    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        raise RuntimeError("Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET in .env")

    all_creds = get_all_credentials(GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET)
    if not all_creds:
        raise RuntimeError("No Google accounts authorized. Run: jimmy auth google")

    seen_emails: set[str] = set()
    count = 0

    for label, creds in all_creds:
        try:
            service = build("gmail", "v1", credentials=creds)

            # Get recent sent + received messages
            for query in ["in:sent", "in:inbox"]:
                msgs = []
                page_token = None
                while len(msgs) < max_messages // 2:
                    resp = service.users().messages().list(
                        userId="me", q=query, maxResults=100, pageToken=page_token
                    ).execute()
                    msgs.extend(resp.get("messages", []))
                    page_token = resp.get("nextPageToken")
                    if not page_token:
                        break

                for msg_stub in msgs[:max_messages // 2]:
                    try:
                        msg = service.users().messages().get(
                            userId="me", id=msg_stub["id"], format="metadata",
                            metadataHeaders=["From", "To", "Cc"]
                        ).execute()
                        headers = {h["name"]: h["value"] for h in msg.get("payload", {}).get("headers", [])}

                        for field in ["From", "To", "Cc"]:
                            val = headers.get(field, "")
                            for name, addr in email.utils.getaddresses([val]):
                                addr = addr.lower().strip()
                                if not addr or addr in seen_emails:
                                    continue
                                # Skip own email, noreply, notifications
                                if any(skip in addr for skip in ["noreply", "no-reply", "notifications", "mailer-daemon", "postmaster", "@columbia.edu" if "columbia" not in label else "XXXNOMATCH"]):
                                    continue
                                seen_emails.add(addr)
                                name = name.strip().strip('"').strip("'")
                                parts = name.split(None, 1) if name else ["", ""]
                                first = parts[0] if parts else ""
                                last = parts[1] if len(parts) > 1 else ""

                                add_contact({
                                    "first_name": first,
                                    "last_name": last,
                                    "email": addr,
                                    "source": f"gmail_{label}",
                                    "affiliations": [],
                                })
                                count += 1
                    except Exception:
                        continue
        except Exception as e:
            print(f"  Gmail {label} failed: {e}")
            continue

    return count


def _parse_location(loc: str) -> tuple[str, str]:
    """Parse location string into (city, region)."""
    if not loc:
        return "", ""
    loc = loc.strip()
    nyc_keywords = {"new york", "nyc", "manhattan", "brooklyn", "queens", "bronx", "staten island"}
    tristate = {"jersey", "nj", "ct", "connecticut", "westchester", "long island", "rumson", "greenwich"}

    low = loc.lower()
    if any(k in low for k in nyc_keywords):
        return loc, "NYC"
    if any(k in low for k in tristate):
        return loc, "Tri-State"
    if "israel" in low or "tel aviv" in low or "jerusalem" in low:
        return loc, "International"
    if any(k in low for k in ("london", "singapore", "hong kong", "tokyo", "paris", "zurich", "geneva")):
        return loc, "International"
    if any(k in low for k in ("los angeles", "la", "san francisco", "sf", "chicago", "boston", "miami", "dc", "washington", "palo alto", "seattle", "dallas", "houston", "atlanta", "denver", "phoenix", "charlotte", "laguna")):
        return loc, "Other-US"
    return loc, ""


def _parse_date(s: str) -> str:
    """Try to parse various date formats into YYYY-MM-DD."""
    if not s:
        return ""
    from datetime import datetime
    for fmt in ("%d %b %Y", "%b %d, %Y", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""
